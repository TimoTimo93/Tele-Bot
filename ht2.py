from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import pytz

class ExcelExporter:
    def __init__(self, activity_tracker):
        self.activity_tracker = activity_tracker
        self.excel_dir = 'excel_exports'
        os.makedirs(self.excel_dir, exist_ok=True)

    def calculate_daily_statistics(self):
        """Tính toán thống kê theo ngày cho mỗi nhân viên"""
        daily_stats = {}
        
        for user_id, username in sorted(self.activity_tracker.usernames.items(), key=lambda x: x[1]):
            user_id_str = str(user_id)
            if user_id_str in self.activity_tracker.daily_work_history:
                work_sessions = self.activity_tracker.daily_work_history[user_id_str]["work_sessions"]
                break_history = self.activity_tracker.daily_work_history[user_id_str].get("break_history", {})
                
                if work_sessions:
                    # Lấy thời gian bắt đầu và kết thúc
                    first_start = datetime.fromisoformat(work_sessions[0]["start_time"])
                    last_end = datetime.fromisoformat(work_sessions[-1]["end_time"])
                    
                    # Tính tổng thời gian
                    total_work_time = timedelta()
                    total_actual_work_time = timedelta()
                    total_break_time = timedelta()
                    
                    for session in work_sessions:
                        work_time = datetime.strptime(session["duration"], '%H:%M:%S')
                        actual_time = datetime.strptime(session["actual_work_time"], '%H:%M:%S')
                        break_time = datetime.strptime(session["break_duration"], '%H:%M:%S')
                        
                        total_work_time += timedelta(hours=work_time.hour, 
                                                   minutes=work_time.minute,
                                                   seconds=work_time.second)
                        total_actual_work_time += timedelta(hours=actual_time.hour,
                                                          minutes=actual_time.minute,
                                                          seconds=actual_time.second)
                        total_break_time += timedelta(hours=break_time.hour,
                                                    minutes=break_time.minute,
                                                    seconds=break_time.second)
                    
                    # Đếm số lần nghỉ các loại
                    break_counts = {
                        "吃饭": 0,
                        "上厕所": 0,
                        "抽烟": 0,
                        "离开": 0
                    }
                    
                    for break_type, breaks in break_history.items():
                        break_counts[break_type] = len(breaks)
                    
                    daily_stats[username] = {
                        "first_start": first_start,
                        "last_end": last_end,
                        "total_work_time": total_work_time,
                        "total_actual_work_time": total_actual_work_time,
                        "total_break_time": total_break_time,
                        "break_counts": break_counts
                    }
                    
        return daily_stats

    def export_statistics(self, filename=None):
        """Xuất thống kê theo định dạng mới"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "考勤统计"

            # Style cho headers
            header_font = Font(bold=True, color="000000", size=12)
            header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers với độ rộng cố định cho từng cột
            headers = [
                "员工姓名",          # Tên nhân viên
                "上班时间",          # Thời gian bắt đầu
                "下班时间",          # Thời gian kết thúc
                "工作总时间",        # Tổng thời gian làm việc
                "实际工作时间",      # Thời gian làm việc thực tế
                "吃饭次数",          # Số lần ăn cơm
                "上厕所次数",        # Số lần đi vệ sinh
                "抽烟次数",          # Số lần hút thuốc
                "离开次数",          # Số lần rời đi
                "休息总时间"         # Tổng thời gian nghỉ
            ]
            
            # Độ rộng cố định cho từng cột (đơn vị là ký tự)
            column_widths = {
                1: 15,  # 姓名
                2: 12,  # 上班时间
                3: 12,  # 下班时间
                4: 12,  # 工作总时间
                5: 12,  # 实际工作时间
                6: 8,   # 吃饭
                7: 8,   # 上厕所
                8: 8,   # 抽烟
                9: 8,   # 离开
                10: 12  # 休息总时间
            }

            # Thiết lập headers với style mới
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', 
                                        vertical='center', 
                                        wrap_text=True)
                cell.border = border
                # Đặt độ rộng cột theo giá trị đã định
                ws.column_dimensions[get_column_letter(col)].width = column_widths[col]

            # Style cho dữ liệu
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal='center', vertical='center')

            # Tính toán và điền dữ liệu
            daily_stats = self.calculate_daily_statistics()
            current_row = 2
            
            # Lấy timezone từ activity_tracker
            timezone = self.activity_tracker.timezone
            tz = pytz.timezone(timezone)
            
            for username, stats in sorted(daily_stats.items()):
                # Chuyển đổi thời gian sang timezone của nhóm
                first_start = stats["first_start"].astimezone(tz)
                last_end = stats["last_end"].astimezone(tz)
                
                # Thông tin cơ bản với thời gian đã chuyển đổi
                ws.cell(row=current_row, column=1).value = username
                ws.cell(row=current_row, column=2).value = first_start.strftime('%H:%M:%S')
                ws.cell(row=current_row, column=3).value = last_end.strftime('%H:%M:%S')
                ws.cell(row=current_row, column=4).value = str(stats["total_work_time"]).split('.')[0]
                ws.cell(row=current_row, column=5).value = str(stats["total_actual_work_time"]).split('.')[0]
                
                # Số lần nghỉ các loại
                ws.cell(row=current_row, column=6).value = stats["break_counts"]["吃饭"]
                ws.cell(row=current_row, column=7).value = stats["break_counts"]["上厕所"]
                ws.cell(row=current_row, column=8).value = stats["break_counts"]["抽烟"]
                ws.cell(row=current_row, column=9).value = stats["break_counts"]["离开"]
                
                # Tổng thời gian nghỉ
                ws.cell(row=current_row, column=10).value = str(stats["total_break_time"]).split('.')[0]
                
                # Áp dụng style cho từng ô dữ liệu
                for col in range(1, 11):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                
                current_row += 1

            # Tạo tên file mặc định nếu không được cung cấp
            if not filename:
                current_date = datetime.now().strftime('%Y%m%d')
                clean_group_name = "".join(c for c in self.activity_tracker.group_name 
                                         if c.isalnum() or c.isspace()).strip()
                filename = os.path.join(self.excel_dir, 
                                      f"{clean_group_name}_statistics_{current_date}.xlsx")

            # Lưu file
            wb.save(filename)
            return True, filename

        except Exception as e:
            print(f"Lỗi khi xuất Excel: {str(e)}")
            return False, str(e)
