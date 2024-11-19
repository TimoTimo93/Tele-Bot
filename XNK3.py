from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from datetime import datetime
import os
import traceback

class ExcelExporter:
    def __init__(self, group_name):
        self.group_name = group_name

    def export_to_excel(self, data, config, date_str=None, force_export=False, create_new=False):
        """Xuất dữ liệu ra file Excel"""
        try:
            if date_str is None:
                date_str = datetime.now().strftime('%Y%m%d')
            
            # Kiểm tra điều kiện xuất
            current_time = datetime.now()
            if not force_export and current_time.hour != 0 and current_time.minute != 0:
                print("Chưa đến thời điểm xuất dữ liệu")
                return None  # Thêm return None ở đây
                
            # Nếu create_new=True, tăng tháng lên 1 để tạo file mới
            if create_new:
                current_date = datetime.strptime(date_str, '%Y%m%d')
                if current_date.month == 12:
                    next_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    next_date = current_date.replace(month=current_date.month + 1)
                date_str = next_date.strftime('%Y%m%d')

            # Tạo border style ngay từ đầu
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Lấy các thông số từ config
            fee_rate = config.get('fee_rate', 0)
            exchange_rate = config.get('exchange_rate', 0)
            currency_type = config.get('currency_type')
            
            # Chỉ lấy tổng tiền giao dịch thực tế của người dùng
            total_in = data.get('user_total_in', 0)  # Tổng tiền người dùng nạp
            total_out = data.get('user_total_out', 0)  # Tổng tiền người dùng rút
            
            # Tính toán các giá trị dựa trên giao dịch thực tế
            expected_out = total_in * (1 - fee_rate/100)
            remaining = expected_out - total_out

            # Thêm các giá trị vào dictionary data
            data['expected_currency'] = expected_out / exchange_rate if exchange_rate > 0 else 0
            data['total_currency_out'] = total_out / exchange_rate if exchange_rate > 0 else 0
            data['remaining_currency'] = remaining / exchange_rate if exchange_rate > 0 else 0
            
            file_path = f"data/{self.group_name}/{self.group_name}每月流水{date_str[0:4]}年{date_str[4:6]}.xlsx"
            
            # Tạo workbook và thiết lập sheet
            wb = Workbook()
            ws = self._setup_summary_sheet(wb, currency_type)
            
            # Thêm dữ liệu ngày hiện tại
            current_date = datetime.now().strftime('%Y/%m/%d')
            ws['A3'] = current_date
            ws['B3'] = fee_rate
            ws['C3'] = exchange_rate
            ws['D3'] = total_in  # Tổng tiền người dùng nạp
            ws['E3'] = expected_out  # Số tiền cần xuất sau phí
            ws['F3'] = data['expected_currency']  # Số tiền cần xuất quy đổi
            ws['G3'] = total_out  # Tổng tiền đã xuất
            ws['H3'] = data['total_currency_out']  # Tổng tiền đã xuất quy đổi
            ws['I3'] = remaining  # Số tiền còn lại
            ws['J3'] = data['remaining_currency']  # Số tiền còn lại quy đổi
            
            # Lưu file
            os.makedirs(f"data/{self.group_name}", exist_ok=True)
            wb.save(file_path)
            return file_path
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            traceback.print_exc()
            return None

    def _setup_summary_sheet(self, wb, currency_type):
        """Thiết lập sheet thống kê ngày và tháng"""
        if "日月统计" in wb.sheetnames:
            wb.remove(wb["日月统计"])
        ws = wb.create_sheet("日月统计", 0)
        
        # Khởi tạo last_row mặc định
        last_row = 3
        
        # Thiết lập chiều rộng cột
        for col in range(1, 18):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Merge cells giữ nguyên
        merge_ranges = [
            'A1:A2', 'B1:B2', 'C1:C2', 'D1:D2',
            'E1:F1', 'G1:H1', 'I1:J1',
            'K1:K2', 'L1:M1', 'N1:O1', 'P1:Q1'
        ]
        
        for cell_range in merge_ranges:
            ws.merge_cells(cell_range)
        
        # Headers và giá trị dòng 2 giữ nguyên
        headers = {
            'A1': '日期', 'B1': '费率', 
            'C1': f"{currency_type}汇率", 'D1': '总入款',
            'E1': '应下发', 'G1': '总下发', 'I1': '未下发',
            'K1': '总月下发',
            'L1': '总月应下发', 'N1': '总月总下发', 'P1': '总月未下发'
        }
        
        row2_values = {
            'E2': "过费率", 'G2': "过费率", 'I2': "过费率",
            'L2': "过费率", 'N2': "过费率", 'P2': "过费率",
            'F2': currency_type, 'H2': currency_type, 'J2': currency_type,
            'M2': currency_type, 'O2': currency_type, 'Q2': currency_type
        }
        
        # Áp dụng headers và row2
        for cells in [headers, row2_values]:
            for cell, value in cells.items():
                ws[cell] = value

        # Xử lý dữ liệu từ Sheet1 và Sheet2
        if 'Sheet1' in wb.sheetnames and 'Sheet2' in wb.sheetnames:
            sheet1 = wb['Sheet1']
            sheet2 = wb['Sheet2']
            last_row = 3
            
            # Duyệt qua từng ngày trong Sheet1
            while sheet1[f'A{last_row}'].value and last_row < 100:
                current_date = sheet1[f'A{last_row}'].value
                
                # Copy dữ liệu cơ bản từ Sheet1
                ws[f'A{last_row}'] = current_date  # Ngày
                ws[f'B{last_row}'] = sheet1[f'B{last_row}'].value  # Phí
                ws[f'C{last_row}'] = sheet1[f'C{last_row}'].value  # Tỷ giá
                
                # Tìm dữ liệu trong Sheet2
                sheet2_row = None
                for row in range(2, sheet2.max_row + 1):
                    if sheet2[f'A{row}'].value == current_date:
                        sheet2_row = row
                        break

                if sheet2_row:
                    # Lấy tổng nhập từ Sheet1 - bỏ phần trừ số dư ngày trước
                    total_in = float(sheet1[f'D{last_row}'].value or 0)
                    
                    # Cột D (总入款): Sử dụng trực tiếp tổng nhập, không trừ số dư
                    ws[f'D{last_row}'] = total_in
                    
                    # Tính phí và tỷ giá
                    fee_rate = float(ws[f'B{last_row}'].value or 0)
                    exchange_rate = float(ws[f'C{last_row}'].value or 0)
                    
                    # Cột E,F (应下发): Tính từ tổng số tiền nhập
                    expected_out = total_in * (1 - fee_rate/100)
                    expected_currency = expected_out / exchange_rate if exchange_rate > 0 else 0
                    
                    ws[f'E{last_row}'] = expected_out
                    ws[f'F{last_row}'] = expected_currency
                    
                    # Cột G,H (总下发): Lấy từ Sheet2 số tiền đã xuất
                    ws[f'G{last_row}'] = sheet2[f'C{sheet2_row}'].value  # VND đã xuất
                    ws[f'H{last_row}'] = sheet2[f'D{sheet2_row}'].value  # USD đã xuất
                    
                    # Cột I,J (未下发): Tính số tiền chưa xuất
                    ws[f'I{last_row}'] = expected_out - float(sheet2[f'C{sheet2_row}'].value or 0)
                    ws[f'J{last_row}'] = expected_currency - float(sheet2[f'D{sheet2_row}'].value or 0)
                
                last_row += 1

            # Sửa lại công thức tổng tháng
            sum_formulas = {
                'K3': '=SUM(D3:D33)', 
                'L3': '=SUM(E3:E33)', 
                'M3': '=SUM(F3:F33)', 
                'N3': '=SUM(G3:G33)', 
                'O3': '=SUM(H3:H33)', 
                'P3': '=SUM(I3:I33)',  
                'Q3': '=SUM(J3:J33)' 
            }

            for cell, formula in sum_formulas.items():
                ws[cell] = Translator(formula, cell).translate_formula()
                ws[cell].number_format = '#,##0.00'

        # Định dạng và style giữ nguyên
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        daily_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        monthly_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        red_font = Font(color="C00000", bold=True)

        # Áp dụng định dạng cho tất cả các dòng
        max_row = max(4, last_row)
        for row in range(1, max_row):
            for col in range(1, 18):
                cell = ws[f"{get_column_letter(col)}{row}"]
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if row <= 2:
                    if col <= 10:
                        cell.fill = daily_fill
                        cell.font = white_font
                    else:
                        cell.fill = monthly_fill
                        cell.font = red_font

        return ws

    def export_daily_balance(self, data_from_xnk4):
        try:
            date_str = datetime.now().strftime('%Y%m%d')
            file_path = f"data/{self.group_name}/{self.group_name}每月流水{date_str[0:4]}年{date_str[4:6]}.xlsx"
            
            # Lấy ngày hiện tại
            current_date = datetime.now().strftime('%Y/%m/%d')
            
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb["日月统计"]
                
                # Tìm dòng của ngày hiện tại và dòng trống tiếp theo
                current_row = None
                next_row = 3
                while ws[f'A{next_row}'].value is not None:
                    if ws[f'A{next_row}'].value == current_date:
                        current_row = next_row
                    next_row += 1
                    
                # Xác định target_row:
                # - Nếu là ngày hiện tại đã có dữ liệu -> ghi đè
                # - Nếu là ngày mới -> thêm vào dòng mới
                if current_row:
                    target_row = current_row
                else:
                    target_row = next_row
            else:
                # Tạo file mới nếu chưa tồn tại
                wb = Workbook()
                ws = self._setup_summary_sheet(wb, data_from_xnk4.get('currency_type'))
                target_row = 3
                
            # Tạo style cho border và alignment
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alignment = Alignment(horizontal='center', vertical='center')
            
            # Thêm dữ liệu và định dạng cho từng cột A-J
            data_mapping = {
                'A': current_date,
                'B': data_from_xnk4.get('fee_rate', 0),
                'C': data_from_xnk4.get('exchange_rate', 0),
                'D': data_from_xnk4.get('total_in', 0),
                'E': data_from_xnk4.get('expected_out', 0),
                'F': data_from_xnk4.get('expected_currency', 0),
                'G': data_from_xnk4.get('total_out', 0),
                'H': data_from_xnk4.get('total_currency_out', 0),
                'I': data_from_xnk4.get('remaining', 0),
                'J': data_from_xnk4.get('remaining_currency', 0)
            }
            
            # Thêm dữ liệu và áp dụng định dạng cho các cột A-J
            for col, value in data_mapping.items():
                cell = ws[f'{col}{target_row}']
                cell.value = value
                cell.border = border
                cell.alignment = alignment
                
                # Định dạng số cho các cột tiền tệ (D-J)
                if col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    cell.number_format = '#,##0'
                    
            # Cập nhật công thức tổng tháng (chỉ ở dòng 3)
            if target_row == 3:
                sum_formulas = {
                    'K3': '=SUM(D3:D33)', 
                    'L3': '=SUM(E3:E33)', 
                    'M3': '=SUM(F3:F33)', 
                    'N3': '=SUM(G3:G33)', 
                    'O3': '=SUM(H3:H33)', 
                    'P3': '=SUM(I3:I33)',  
                    'Q3': '=SUM(J3:J33)' 
                }

                for cell, formula in sum_formulas.items():
                    ws[cell] = formula
                    ws[cell].number_format = '#,##0.00'
            
            # Lưu file
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb.save(file_path)
            
            return file_path, data_from_xnk4.get('remaining', 0)
            
        except Exception as e:
            print(f"Error in export_daily_balance: {str(e)}")
            traceback.print_exc()
            return None, None


