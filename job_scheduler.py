from datetime import datetime, time
import pytz
import os
from hethong import ActivityTracker
from XNK import XuatNhapKhoan
from group_config import GroupConfig
import traceback
from telegram.ext import ContextTypes
from ht2 import ExcelExporter as AttendanceExporter
from telegram.error import Forbidden, ChatMigrated
import telegram
from XNK4 import DetailedTransactionHistory
from XNK3 import ExcelExporter as MonthlyExporter
from openpyxl import load_workbook

class JobScheduler:
    def __init__(self):
        self.group_config = GroupConfig()
        self.jobs = {}
        self.schedules = {}
        
    async def start_check_membership(self, context: ContextTypes.DEFAULT_TYPE):
        """Bắt đầu kiểm tra tư cách thành viên"""
        # Kiểm tra ngay lập tức khi khởi động
        await self._check_all_groups_membership(context)
        
        # Lấy tất cả cài đặt nhóm
        settings = self.group_config.get_all_settings()
        
        # Thiết lập kiểm tra hàng ngày cho từng nhóm theo múi giờ riêng
        for chat_id, group_settings in settings.items():
            timezone = group_settings.get('timezone', 'UTC')
            tz = pytz.timezone(timezone)
            job_time = time(hour=23, minute=59, tzinfo=tz)
            
            context.job_queue.run_daily(
                self._check_all_groups_membership,
                time=job_time,
                name=f'check_membership_{chat_id}'
            )

    async def _check_all_groups_membership(self, context: ContextTypes.DEFAULT_TYPE):
        """Kiểm tra tất cả các nhóm"""
        print("Đang kiểm tra tư cách thành viên trong các nhóm...")
        settings = self.group_config.get_all_settings()
        
        total_groups = len(settings)
        checked_groups = 0
        removed_groups = 0
        
        for chat_id in list(settings.keys()):
            try:
                if not await self._check_bot_in_group(context, int(chat_id)):
                    print(f"Bot không còn trong nhóm {chat_id}, đã xóa cấu hình")
                    removed_groups += 1
                checked_groups += 1
                print(f"Đã kiểm tra {checked_groups}/{total_groups} nhóm")
            except Exception as e:
                print(f"Lỗi khi kiểm tra nhóm {chat_id}: {str(e)}")
                checked_groups += 1

        print(f"""
Kết quả kiểm tra:
- Tổng số nhóm: {total_groups}
- Số nhóm đã kiểm tra: {checked_groups}
- Số nhóm đã xóa: {removed_groups}
- Số nhóm còn lại: {total_groups - removed_groups}
""")

    async def send_monthly_xnk_report(self, context, chat_id, chat_title):
        """Gửi báo cáo XNK hàng tháng"""
        try:
            tracker = ActivityTracker(chat_title)
            tz = pytz.timezone(tracker.timezone)
            current_date = datetime.now(tz).strftime('%Y%m%d')
            
            xnk = XuatNhapKhoan(chat_title)
            excel_file = xnk.export_to_excel(current_date)
            
            if excel_file and os.path.exists(excel_file):
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=open(excel_file, 'rb'),
                    caption=f"出入款统计 - {chat_title} - {current_date}"
                )
                
        except Exception as e:
            print(f"Error in send_monthly_xnk_report: {str(e)}")

    def schedule_daily_statistics(self, context, chat_id, chat_title, time_str, timezone='UTC'):
        """Lên lịch gửi thống kê hàng ngày"""
        try:
            hour, minute = map(int, time_str.split(':'))
            tz = pytz.timezone(timezone)
            job_time = time(hour=hour, minute=minute, tzinfo=tz)
            
            job_data = {
                'chat_id': chat_id,
                'chat_title': chat_title,
                'timezone': timezone
            }
            
            # Xóa job cũ nếu có
            job_name = f"daily_stats_{chat_id}"
            self.remove_job(job_name, context)
            
            # Tạo job mới sử dụng _send_attendance_report
            job = context.job_queue.run_daily(
                self._send_attendance_report,  # Sử dụng callback mới
                job_time,
                data=job_data,
                name=job_name
            )
            
            # Lưu job vào dictionary
            if str(chat_id) not in self.jobs:
                self.jobs[str(chat_id)] = {}
            self.jobs[str(chat_id)][job_name] = job
            
            return True, f"✅ 已设置每日自动发送时间为 {time_str}"
            
        except Exception as e:
            print(f"Error scheduling daily statistics: {str(e)}")
            return False, "❌ 设置自动发送时间失败"

    async def send_monthly_report(self, context: ContextTypes.DEFAULT_TYPE):
        """Callback function cho job tự động gửi báo cáo tháng"""
        job = context.job
        chat_id = job.data['chat_id']
        try:
            chat = await context.bot.get_chat(chat_id)
            
            # Khởi tạo XuatNhapKhoan và xuất báo cáo
            xnk = XuatNhapKhoan(chat.title)
            excel_file = xnk.export_to_excel()
            
            if excel_file and os.path.exists(excel_file):
                await context.bot.send_document(
                    chat_id=chat_id,
                    
                    document=open(excel_file, 'rb'),
                    caption=f"出入款统计 - {chat.title}"
                )
        except Exception as e:
            print(f"Error in automatic send_monthly_report: {str(e)}")

    def _remove_existing_jobs(self, chat_id: str, job_type: str):
        """Xóa các jobs hiện có của một chat_id và job_type cụ thể"""
        try:
            if chat_id in self.jobs:
                jobs_to_remove = []
                for job_id, job in self.jobs[chat_id].items():
                    if job_id.startswith(job_type):
                        # Chỉ cần gọi schedule_removal() để xóa job
                        job.schedule_removal()
                        jobs_to_remove.append(job_id)
                
                # Xóa jobs khỏi dictionary
                for job_id in jobs_to_remove:
                    del self.jobs[chat_id][job_id]
                
        except Exception as e:
            print(f"Error removing jobs: {str(e)}")

    async def schedule_monthly_report(self, context, chat_id, chat_title, day, time_str, timezone='UTC'):
        try:
            day = int(day)
            if not (1 <= day <= 31):
                return False, "❌ 日期必须在1-31之间"

            self._remove_existing_jobs(str(chat_id), 'monthly_report')
            
            # Parse time
            hour, minute = map(int, time_str.split(':'))
            tz = pytz.timezone(timezone)
            job_time = time(hour=hour, minute=minute, tzinfo=tz)
            
            job_data = {
                'chat_id': chat_id,
                'chat_title': chat_title,
                'day': day,
                'timezone': timezone,
                'last_run': None  # Thêm trường này để theo dõi lần chạy cuối
            }
            
            # Tạo job mới
            job = context.job_queue.run_monthly(
                callback=self._send_monthly_report,
                when=job_time,
                day=day,
                data=job_data,
                name=f'monthly_report_{chat_id}'
            )
            
            if str(chat_id) not in self.jobs:
                self.jobs[str(chat_id)] = {}
            self.jobs[str(chat_id)][f'monthly_report_{day}_{time_str}'] = job

            return True, f"✅ 已设置每月{day}日 {time_str} 自动发送报告"
            
        except Exception as e:
            print(f"Error scheduling monthly report: {str(e)}")
            traceback.print_exc()
            return False, "❌ 设置自动发送时间失败"

    async def _send_monthly_report(self, context: ContextTypes.DEFAULT_TYPE):
        try:
            job = context.job
            chat_id = job.data['chat_id']
            chat_title = job.data['chat_title']
            timezone = job.data.get('timezone', 'UTC')
            tz = pytz.timezone(timezone)
            current_time = datetime.now(tz)
            
            # Kiểm tra xem báo cáo đã được gửi trong tháng này chưa
            last_run = job.data.get('last_run')
            if last_run:
                last_run = datetime.strptime(last_run, '%Y-%m-%d').replace(tzinfo=tz)
                if last_run.year == current_time.year and last_run.month == current_time.month:
                    print(f"Monthly report already sent this month for {chat_title}")
                    return
            
            # Cập nhật last_run
            job.data['last_run'] = current_time.strftime('%Y-%m-%d')
            
            # Tiếp tục với logic gửi báo cáo hiện tại
            print(f"Starting monthly report generation from XNK3 for {chat_title}")
            
            if not await self._check_bot_in_group(context, chat_id):
                return
            
            try:
                # Tạo đối tượng DetailedTransactionHistory để lấy config và data
                daily_report = DetailedTransactionHistory(chat_title)
                config = daily_report._load_config()
                data = daily_report._load_data()
                
                # Tính toán các giá trị cần thiết
                total_in = sum(t['amount'] for t in data['transactions'] if t['type'] == '入款')
                total_out = sum(t['amount'] for t in data['transactions'] if t['type'] == '下发')
                
                fee_rate = config.get('fee_rate', 0)
                exchange_rate = config.get('exchange_rate', 0)
                currency_type = config.get('currency_type', '')
                
                expected_out = total_in * (1 - fee_rate/100)
                remaining = expected_out - total_out
                
                # Chuẩn bị dữ liệu cho XNK3
                data = {
                    'fee_rate': fee_rate,
                    'exchange_rate': exchange_rate,
                    'currency_type': currency_type,
                    'user_total_in': total_in,
                    'user_total_out': total_out,
                    'expected_out': expected_out,
                    'expected_currency': expected_out / exchange_rate if exchange_rate > 0 else 0,
                    'total_currency_out': total_out / exchange_rate if exchange_rate > 0 else 0,
                    'remaining': remaining,
                    'remaining_currency': remaining / exchange_rate if exchange_rate > 0 else 0
                }
                
                print(f"Data to be exported: {data}")  # Debug log
                
                # Khởi tạo MonthlyExporter
                excel_exporter = MonthlyExporter(chat_title)
                timezone = job.data.get('timezone', 'UTC')
                tz = pytz.timezone(timezone)
                current_date = datetime.now(tz).strftime('%Y%m%d')
                
                # Thêm force_export=True để đảm bảo xuất báo cáo
                excel_file = excel_exporter.export_to_excel(data, config, date_str=current_date, force_export=True)
                
                # Nếu file được tạo thành công, mở lại để thêm công thức SUM
                if excel_file and os.path.exists(excel_file):
                    wb = load_workbook(excel_file)
                    ws = wb["日月统计"]
                    
                    # Thêm công thức SUM
                    sum_formulas = {
                        'K3': '=SUM(D3:D33)', 
                        'L3': '=SUM(E3:E33)', 
                        'M3': '=SUM(F3:F33)', 
                        'N3': '=SUM(G3:G33)', 
                        'O3': '=SUM(H3:H33)', 
                        'P3': '=SUM(I3:I33)',  
                        'Q3': '=SUM(J3:J33)' 
                    }
                    
                    for cell, formula in sum_formulas.items():
                        ws[cell] = formula
                        ws[cell].number_format = '#,##0.00'
                    
                    # Lưu lại file
                    wb.save(excel_file)
                    
                    print(f"Sending XNK3 monthly report: {excel_file}")
                    await context.bot.send_document(
                        chat_id=chat_id,
                        document=open(excel_file, 'rb'),
                        caption=f"月度统计报告 - {chat_title}"
                    )
                else:
                    print(f"Failed to generate XNK3 monthly report for {chat_title}")
            except Exception as e:
                print(f"Error generating XNK3 monthly report: {str(e)}")
                traceback.print_exc()
                
        except Exception as e:
            print(f"Error in _send_monthly_report: {str(e)}")
            traceback.print_exc()

    async def clear_daily_transactions(self, context, chat_id, chat_title):
        try:
            print(f"Starting daily auto-clear for {chat_title}")
            
            # Kiểm tra và cập nhật chat_id nếu group đã được nâng cấp
            try:
                chat = await context.bot.get_chat(chat_id)
            except ChatMigrated as e:
                new_chat_id = e.new_chat_id
                print(f"Group {chat_title} migrated to supergroup. Updating chat_id from {chat_id} to {new_chat_id}")
                
                # Cập nhật chat_id trong group_config
                settings = self.group_config.get_group_settings(str(chat_id))
                if settings:
                    settings['chat_id'] = str(new_chat_id)
                    self.group_config.set_group_settings(str(new_chat_id), chat_title, settings)
                    # Xóa cấu hình cũ
                    self.group_config.remove_group_settings(str(chat_id))
                
                # Cập nhật chat_id cho các jobs
                self._update_jobs_chat_id(context, str(chat_id), str(new_chat_id))
                
                # Sử dụng chat_id mới
                chat_id = new_chat_id
                chat = await context.bot.get_chat(chat_id)
            
            tracker = ActivityTracker(chat_title)
            tz = pytz.timezone(tracker.timezone)
            
            # 1. Tạo báo cáo ngày trước khi xóa dữ liệu
            daily_report = DetailedTransactionHistory(chat_title)
            report_file = daily_report.create_daily_report()
            
            if report_file and os.path.exists(report_file):
                # Gửi file báo cáo vào nhóm
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=open(report_file, 'rb'),
                    caption=f"每日报告 - {chat_title}"
                )
            
            # 2. Sau khi gửi báo cáo, thực hiện xóa dữ liệu
            xnk = XuatNhapKhoan(chat_title)
            message, carried_amount = xnk.auto_clear_daily()
            
            # 3. Gửi thông báo kết quả kèm số dư
            await context.bot.send_message(
                chat_id=chat_id,
                text=(
                    f"<b>🔄 每日自动结算完成</b>\n\n"
                    f"{message}\n"
                ),
                parse_mode='HTML'
            )
        except Exception as e:
            print(f"Error in clear_daily_transactions: {str(e)}")
            traceback.print_exc()

    def _update_jobs_chat_id(self, context, old_chat_id: str, new_chat_id: str):
        """Cập nhật chat_id cho tất cả các jobs của một group"""
        try:
            if old_chat_id in self.jobs:
                # Tạo jobs mới với chat_id mới
                for job_name, old_job in self.jobs[old_chat_id].items():
                    job_data = old_job.data
                    job_data['chat_id'] = int(new_chat_id)  # Cập nhật chat_id trong job_data
                    
                    # Tạo job mới với chat_id đã cập nhật
                    new_job = context.job_queue.scheduler.reschedule_job(
                        old_job.name,
                        trigger=old_job.trigger,
                        kwargs={'job': {'data': job_data}}
                    )
                    
                    # Cập nhật dictionary jobs
                    if new_chat_id not in self.jobs:
                        self.jobs[new_chat_id] = {}
                    self.jobs[new_chat_id][job_name] = new_job
                
                # Xóa jobs cũ
                del self.jobs[old_chat_id]
                
        except Exception as e:
            print(f"Error updating jobs chat_id: {str(e)}")
            traceback.print_exc()

    def schedule_daily_clear(self, context, chat_id, chat_title):
        """Lập lịch xóa transactions.json hàng ngày vào 00:00"""
        try:
            # Xóa job cũ nếu có
            job_name = f"daily_clear_{chat_id}"
            current_jobs = context.job_queue.get_jobs_by_name(job_name)
            for job in current_jobs:
                job.schedule_removal()
            
            # Lấy timezone từ cài đặt nhóm
            settings = self.group_config.get_group_settings(str(chat_id))
            tz = pytz.timezone(settings.get('timezone', 'UTC'))
            
            # Tạo job mới chạy vào 00:00
            context.job_queue.run_daily(
                callback=lambda ctx: self.clear_daily_transactions(ctx, chat_id, chat_title),
                time=time(hour=10, minute=7, tzinfo=tz),
                name=job_name
            )
            
            return True
            
        except Exception as e:
            print(f"Error scheduling daily clear: {str(e)}")
            return False

    async def load_all_schedules(self, context):
        """Khôi phục tất cả lịch khi bot khởi động"""
        try:
            # Bắt đầu job kiểm tra định kỳ - sử dụng await
            await self.start_check_membership(context)
            
            config = self.group_config.get_all_settings()
            for chat_id, settings in config.items():
                # Khôi phục lịch gửi hàng ngày
                if 'auto_send_time' in settings:
                    self.schedule_daily_statistics(
                        context,
                        int(chat_id),
                        settings['title'],
                        settings['auto_send_time'],
                        settings['timezone']
                    )
                
                # Khôi phục lịch gửi hàng tháng
                if 'monthly_report_day' in settings and 'monthly_report_time' in settings:
                    await self.schedule_monthly_report(  # Thêm await
                        context,
                        int(chat_id),
                        settings['title'],
                        settings['monthly_report_day'],
                        settings['monthly_report_time'],
                        settings['timezone']
                    )
                
                # Thêm lịch xóa transactions.json hàng ngày
                if 'timezone' in settings:
                    self.schedule_daily_clear(
                        context,
                        int(chat_id),
                        settings['title']
                    )
                    
        except Exception as e:
            print(f"Error loading schedules: {str(e)}")

    def remove_job(self, job_name, context):
        """Xóa job theo tên"""
        try:
            current_jobs = context.job_queue.get_jobs_by_name(job_name)
            for job in current_jobs:
                job.schedule_removal()
            return True
        except Exception as e:
            print(f"Error removing job {job_name}: {str(e)}")
            return False

    async def schedule_attendance_report(self, context, chat_id, chat_title, time_str):
        try:
            self._remove_existing_jobs(str(chat_id), 'attendance_report')
            
            hour, minute = map(int, time_str.split(':'))
            job_time = time(hour=hour, minute=minute)
            
            job_data = {
                'chat_id': chat_id,
                'chat_title': chat_title,
                'report_type': 'attendance'  # Đánh dấu đây là báo cáo chấm công
            }
            
            job = context.job_queue.run_daily(
                self._send_attendance_report,
                job_time,
                data=job_data
            )
            
            if str(chat_id) not in self.jobs:
                self.jobs[str(chat_id)] = {}
            self.jobs[str(chat_id)][f'attendance_report_{time_str}'] = job
            
            return True, f"已设置每天 {time_str} 自动发送考勤统计"
            
        except Exception as e:
            print(f"Error scheduling attendance report: {str(e)}")
            return False, "设置自动发送时间失败"

    async def _send_attendance_report(self, context: ContextTypes.DEFAULT_TYPE):
        """Callback function cho job tự động gửi thống kê"""
        try:
            chat_id = context.job.data['chat_id']
            
            # Kiểm tra bot còn trong nhóm không
            if not await self._check_bot_in_group(context, chat_id):
                return
            
            chat = await context.bot.get_chat(chat_id)
            
            # Tạo thư mục nếu chưa tồn tại
            os.makedirs('statistics', exist_ok=True)
            
            # Khởi tạo ActivityTracker trước
            tracker = ActivityTracker(chat.title)
            exporter = AttendanceExporter(tracker)  # Truyền tracker vào
            
            success, excel_file = exporter.export_statistics()
            
            if success and excel_file and os.path.exists(excel_file):
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=open(excel_file, 'rb'),
                    caption=f"考勤统计 - {chat.title}"
                )
                
        except Forbidden as e:
            print(f"Bot đã bị kick khỏi nhóm {chat_id}, hủy job tự động")
            self._remove_schedule(chat_id)

    def _remove_schedule(self, chat_id: int):
        """Hủy tất cả job tự động cho một nhóm"""
        try:
            str_chat_id = str(chat_id)
            if str_chat_id in self.jobs:
                # Xóa tất cả jobs của nhóm
                for job_name, job in self.jobs[str_chat_id].items():
                    try:
                        job.schedule_removal()
                    except Exception as e:
                        print(f"Lỗi khi xóa job {job_name}: {str(e)}")
                
                # Xóa khỏi dictionary
                del self.jobs[str_chat_id]
                
                # Xóa cấu hình nhóm
                self.group_config.remove_group_settings(str_chat_id)
                
                print(f"Đã hủy tất cả job tự động cho nhóm {chat_id}")
                
        except Exception as e:
            print(f"Lỗi khi hủy job cho nhóm {chat_id}: {str(e)}")

    def list_scheduled_jobs(self):
        """In ra tất cả jobs đã lên lịch"""
        for chat_id, jobs in self.jobs.items():
            print(f"\nJobs for chat {chat_id}:")
            for job_id, job in jobs.items():
                next_run = job.next_t if hasattr(job, 'next_t') else 'Unknown'
                print(f"- {job_id}: Next run at {next_run}")

    def _get_current_time(self, timezone='UTC'):
        """Lấy thời gian hiện tại theo timezone"""
        try:
            tz = pytz.timezone(timezone)
            return datetime.now(tz)
        except Exception as e:
            print(f"Error getting current time: {str(e)}")
            return datetime.now(pytz.UTC)

    async def _check_bot_in_group(self, context: ContextTypes.DEFAULT_TYPE, chat_id: int) -> bool:
        """Kiểm tra xem bot còn trong nhóm không"""
        try:
            await context.bot.get_chat_member(chat_id, context.bot.id)
            return True
        except (Forbidden, telegram.error.BadRequest) as e:
            print(f"Bot đã bị kick khỏi nhóm {chat_id}, hủy job tự động")
            self._remove_schedule(chat_id)
            self.group_config.remove_group_settings(str(chat_id))
            return False