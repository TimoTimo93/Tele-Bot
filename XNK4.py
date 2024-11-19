import pandas as pd
from datetime import datetime
import pytz
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
import traceback
from XNK3 import ExcelExporter

class DetailedTransactionHistory:
    def __init__(self, group_name, chat_id=None):
        self.group_name = group_name
        self.chat_id = chat_id
        self.data_file = f"data/{group_name}/transactions.json"
        self.config_file = f"data/{group_name}/config.json"
        self.group_settings_file = "group_settings.json"  # File này nằm ở thư mục gốc
        
    def _load_data(self):
        with open(self.data_file, 'r', encoding='utf-8') as f:
            return json.load(f)
            
    def _load_config(self):
        with open(self.config_file, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_group_timezone(self):
        """Lấy timezone của nhóm từ group_settings"""
        try:
            # Tạo file group_settings.json nếu chưa tồn tại
            if not os.path.exists(self.group_settings_file):
                os.makedirs(os.path.dirname(self.group_settings_file), exist_ok=True)
                with open(self.group_settings_file, 'w', encoding='utf-8') as f:
                    json.dump({self.group_name: {"timezone": "UTC"}}, f, ensure_ascii=False, indent=2)
            
            with open(self.group_settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                group_settings = settings.get(self.group_name, {})
                return group_settings.get('timezone', 'UTC')
        except Exception as e:
            print(f"Error loading group timezone: {str(e)}")
            return 'UTC'

    def _create_header_style(self, color):
        return {
            'fill': PatternFill(start_color=color, end_color=color, fill_type='solid'),
            'font': Font(bold=True, color='FFFFFF', name='.VnArial', size=20),  # Tăng font size lên 20
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def create_daily_report(self, date_str=None):
        try:
            print("Starting create daily report...")  # Debug log
            data = self._load_data()
            config = self._load_config()
            timezone = self._load_group_timezone()
            
            # Load các giá trị config ngay từ đầu
            currency_type = config.get('currency_type', '')
            fee_rate = config.get('fee_rate', 0)
            exchange_rate = config.get('exchange_rate', 0)
            
            if date_str is None:
                current_time = datetime.now(pytz.timezone(timezone))
                date_str = current_time.strftime('%Y%m%d')
            
            print(f"Creating report for date: {date_str}")  # Debug log
            
            wb = Workbook()
            ws = wb.active
            ws.title = "每日报告"

            # Tắt gridlines
            ws.sheet_view.showGridLines = False

            # Thiết lập font mặc định cho toàn bộ worksheet với size lớn hơn
            default_font = Font(name='.VnArial', size=12)  # Tăng font size
            for row in ws.rows:
                for cell in row:
                    cell.font = default_font

            # Định nghĩa styles với font size lớn hơn cho headers
            header_styles = {
                '入款数': self._create_header_style('4472C4'),
                '下发数': self._create_header_style('70AD47'), 
                '总统计数据': self._create_header_style('ED7D31'),
                '客户统计': self._create_header_style('7030A0'),
                '操作员统计': self._create_header_style('C00000')
            }

            # Điều chỉnh độ cao của các dòng tiêu đề
            ws.row_dimensions[1].height = 40  # Dòng tiêu đề chính
            ws.row_dimensions[2].height = 25  # Dòng tiêu đề phụ

            # Điều chỉnh độ rộng cột
            column_widths = {
                'A': 15,  # 时间
                'B': 15,  # 金额
                'C': 15,  # 操作昵称 (người thao tác)
                'D': 15,  # 回复昵称 (người gửi tin nhắn)
                'E': 15,
                'F': 15,
                'G': 15,
                'H': 15,
                'I': 15,
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Hàm helper để áp dụng border và căn giữa với font size lớn hơn
            def apply_cell_format(cell, has_border=True, is_header=False):
                if is_header:
                    cell.font = Font(name='.VnArial', size=12, bold=True)
                else:
                    cell.font = Font(name='.VnArial', size=12)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if has_border:
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            # Sửa lại hàm apply_style để có font size lớn hơn
            def apply_style(cell, style_dict):
                cell.fill = style_dict['fill']
                cell.font = Font(bold=True, color='FFFFFF', name='.VnArial', size=20)  # Tăng font size lên 20
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = style_dict['border']

            # 1. Phần nhập khoản (入款数) - Bắt đầu từ cột A
            deposits = [t for t in data['transactions'] if t['type'] == '入款']
            deposits_count = len(deposits)
            ws.merge_cells('A1:D1')
            ws['A1'] = f'入款数 ({deposits_count})'
            for cell in ws['A1:D1'][0]:
                apply_style(cell, header_styles['入款数'])

            # Headers cho nhập khoản
            headers = ['时间', '金额', '操作昵称', '回复昵称']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                apply_cell_format(cell, is_header=True)

            # Thêm dữ liệu nhập khoản
            current_row = 3
            for trans in deposits:
                formatted_time = trans['timestamp']
                original_sender = trans.get('original_sender', '')
                print(f"Processing deposit - Original sender: {original_sender}")  # Debug log
                
                cells = [
                    (current_row, 1, formatted_time),
                    (current_row, 2, trans['amount']),
                    (current_row, 3, trans['username']),
                    (current_row, 4, original_sender),  # 回复昵称
                ]
                for row, col, value in cells:
                    cell = ws.cell(row=row, column=col, value=value)
                    apply_cell_format(cell)
                current_row += 1

            # 2. Phần xuất khoản (下发数) - Bắt đầu từ cột F
            withdrawals = [t for t in data['transactions'] if t['type'] == '下发']
            withdrawals_count = len(withdrawals)
            ws.merge_cells('F1:I1')
            ws['F1'] = f'下发数 ({withdrawals_count})'
            for cell in ws['F1:I1'][0]:
                apply_style(cell, header_styles['下发数'])

            # Headers cho xuất khoản
            for col, header in enumerate(headers, 6):
                cell = ws.cell(row=2, column=col, value=header)
                apply_cell_format(cell, is_header=True)

            # Thêm dữ liệu xuất khoản
            withdrawal_row = 3
            for trans in withdrawals:
                formatted_time = trans['timestamp']
                original_sender = trans.get('original_sender', '')
                print(f"Processing withdrawal - Original sender: {original_sender}")  # Debug log
                
                cells = [
                    (withdrawal_row, 6, formatted_time),
                    (withdrawal_row, 7, trans['amount']),
                    (withdrawal_row, 8, trans['username']),
                    (withdrawal_row, 9, original_sender),  # 回复昵称
                ]
                for row, col, value in cells:
                    cell = ws.cell(row=row, column=col, value=value)
                    apply_cell_format(cell)
                withdrawal_row += 1

            # Tìm dòng cuối cùng của dữ liệu giao dịch
            last_transaction_row = max(current_row, withdrawal_row)

            # 3. Phần thống kê khách hàng
            customer_row = last_transaction_row + 2
            ws.merge_cells(f'A{customer_row}:E{customer_row}') 
            ws[f'A{customer_row}'] = '客户统计(回复人分类)'
            for cell in ws[f'A{customer_row}:E{customer_row}'][0]:  # Áp dụng style cho tất cả các ô từ A đến I
                apply_style(cell, header_styles['客户统计'])
            ws.row_dimensions[customer_row].height = 35  # Dòng tiêu đề chính
            ws.row_dimensions[customer_row + 1].height = 25  # Dòng tiêu đề phụ

            customer_row += 1
            customer_headers = ['回复人账号', '回复人昵称', '总单数', '总金额', '换算']
            for col, header in enumerate(customer_headers, 1):
                cell = ws.cell(row=customer_row, column=col, value=header)
                apply_cell_format(cell, is_header=True)

            # Hiển thị thống kê khách hàng
            user_stats = {}
            for trans in data['transactions']:
                username = trans['username']
                original_sender = trans.get('original_sender', '')
                amount = trans['amount']

                if username not in user_stats:
                    user_stats[username] = {
                        'with_sender': {'count': 0, 'amount': 0, 'original_sender': ''},
                        'without_sender': {'count': 0, 'amount': 0}
                    }

                if original_sender:
                    user_stats[username]['with_sender']['count'] += 1
                    user_stats[username]['with_sender']['amount'] += amount
                    user_stats[username]['with_sender']['original_sender'] = original_sender
                else:
                    user_stats[username]['without_sender']['count'] += 1
                    user_stats[username]['without_sender']['amount'] += amount

            for username, stats in user_stats.items():
                if stats['with_sender']['count'] > 0:
                    customer_row += 1
                    cells = [
                        (customer_row, 1, username),
                        (customer_row, 2, stats['with_sender']['original_sender']),
                        (customer_row, 3, stats['with_sender']['count']),
                        (customer_row, 4, stats['with_sender']['amount']),
                        (customer_row, 5, currency_type)
                    ]
                    for row, col, value in cells:
                        cell = ws.cell(row=row, column=col, value=value)
                        apply_cell_format(cell)

                if stats['without_sender']['count'] > 0:
                    customer_row += 1
                    cells = [
                        (customer_row, 1, username),
                        (customer_row, 2, "无回复人"),
                        (customer_row, 3, stats['without_sender']['count']),
                        (customer_row, 4, stats['without_sender']['amount']),
                        (customer_row, 5, currency_type)
                    ]
                    for row, col, value in cells:
                        cell = ws.cell(row=row, column=col, value=value)
                        apply_cell_format(cell)

            # Tính toán operator_stats trước
            operator_stats = {}
            for trans in data['transactions']:
                # Lấy operator từ trường operator, nếu không có thì dùng username
                operator = trans.get('operator', trans.get('username', ''))
                print(f"Processing transaction: {trans}")  # Debug log
                print(f"Operator: {operator}")  # Debug log
                
                if operator:  # Chỉ xử lý khi có thông tin người thao tác
                    if operator not in operator_stats:
                        operator_stats[operator] = {
                            'in_count': 0,
                            'out_count': 0,
                            'in_amount': 0,
                            'out_amount': 0
                        }
                    
                    # Phân biệt giữa nhập khoản và xuất khoản
                    if trans['type'] == '入款':
                        operator_stats[operator]['in_count'] += 1
                        operator_stats[operator]['in_amount'] += trans['amount']
                    elif trans['type'] == '下发':
                        operator_stats[operator]['out_count'] += 1
                        operator_stats[operator]['out_amount'] += trans['amount']

            print(f"Final operator_stats: {operator_stats}")  # Debug log

            # Sau đó mới vẽ bảng thống kê người thao tác
            operator_row = customer_row + len(user_stats.items()) + 2
            ws.merge_cells(f'A{operator_row}:E{operator_row}')  # Merge từ A đến I
            ws[f'A{operator_row}'] = '每个人的总统计'
            for cell in ws[f'A{operator_row}:E{operator_row}'][0]:  # Áp dụng style cho tất cả các ô
                apply_style(cell, header_styles['操作员统计'])
            ws.row_dimensions[operator_row].height = 35  # Dòng tiêu đề chính
            ws.row_dimensions[operator_row + 1].height = 25  # Dòng tiêu đề phụ

            operator_row += 1
            operator_headers = ['操作昵称', '入款单数', '入款金额', '下发单数', '下发金额']
            for col, header in enumerate(operator_headers, 1):
                cell = ws.cell(row=operator_row, column=col, value=header)
                apply_cell_format(cell, is_header=True)

            # Thêm dữ liệu thống kê người thao tác
            for operator, stats in operator_stats.items():
                operator_row += 1
                cells = [
                    (operator_row, 1, operator),  # 操作昵称 - Tên người thao tác
                    (operator_row, 2, stats['in_count']),  # 入款单数 - Số đơn nhập khoản
                    (operator_row, 3, stats['in_amount']),  # 入款金额 - Tổng tiền nhập khoản
                    (operator_row, 4, stats['out_count']),  # 下发单数 - Số đơn xuất khoản
                    (operator_row, 5, stats['out_amount'])  # 下发金额 - Tổng tiền xuất khoản
                ]
                for row, col, value in cells:
                    cell = ws.cell(row=row, column=col, value=value)
                    apply_cell_format(cell)

            # 4. Phần thống kê tổng - Di chuyển về cột A và điều chỉnh độ rộng
            stats_row = operator_row + len(operator_stats) + 2
            ws.merge_cells(f'A{stats_row}:I{stats_row}')  # Điều chỉnh merge cells giống 入款数
            ws[f'A{stats_row}'] = '总统计数据'
            for cell in ws[f'A{stats_row}:I{stats_row}'][0]:
                apply_style(cell, header_styles['总统计数据'])
            ws.row_dimensions[stats_row].height = 35  # Dòng tiêu đề chính
            ws.row_dimensions[stats_row + 1].height = 25  # Dòng tiêu đề phụ
            # Thêm dữ liệu thống kê tổng
            stats_row += 1
            total_in = data.get('total_in', 0)
            fee_rate = config.get('fee_rate', 0)
            currency_type = config.get('currency_type', '')
            exchange_rate = config.get('exchange_rate', 0)
            total_out = data.get('total_out', 0)
            expected_out = total_in * (1 - fee_rate/100)
            remaining = expected_out - total_out

            # Tính giá trị quy đổi
            expected_currency = expected_out / exchange_rate if exchange_rate > 0 else 0
            total_currency_out = total_out / exchange_rate if exchange_rate > 0 else 0
            remaining_currency = remaining / exchange_rate if exchange_rate > 0 else 0

            # Sửa lại thứ tự các headers và values, thêm cột quy đổi
            stats_headers = [
                '费率', 
                f'{currency_type}汇率', 
                '总入款',
                '应下发', f'应下发{currency_type}',
                '总下发', f'总下发{currency_type}',
                '未下发', f'未下发{currency_type}'
            ]
            stats_values = [
                f"{fee_rate}%",  # Thay đổi cách hiển thị phí
                exchange_rate,
                total_in,
                expected_out, expected_currency,
                total_out, total_currency_out,
                remaining, remaining_currency
            ]

            for col, (header, value) in enumerate(zip(stats_headers, stats_values), 1):
                header_cell = ws.cell(row=stats_row, column=col, value=header)
                value_cell = ws.cell(row=stats_row + 1, column=col, value=value)
                
                # Áp dụng định dạng font cho header
                header_cell.font = Font(name='.VnArial', size=12, bold=True)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Áp dụng định dạng cho value cell
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Định dạng số đặc biệt cho tỷ giá và các cột tiền tệ
                if col == 2:  # Cột tỷ giá
                    value_cell.number_format = '0.00'  # Hiển thị 2 số thập phân
                elif col >= 3:  # Các cột tiền tệ
                    value_cell.number_format = '#,##0'

            # Tính tổng số tiền nhập/xuất
            total_in = sum(t['amount'] for t in data['transactions'] if t['type'] == '入款')
            total_out = sum(t['amount'] for t in data['transactions'] if t['type'] == '下发')
            data['total_in'] = total_in
            data['total_out'] = total_out

            # Chuẩn bị dữ liệu để chuyển sang XNK3
            data_for_xnk3 = {
                'fee_rate': fee_rate,
                'exchange_rate': exchange_rate,
                'currency_type': currency_type,
                'total_in': total_in,
                'expected_out': expected_out,
                'expected_currency': expected_currency,
                'total_out': total_out,
                'total_currency_out': total_currency_out,
                'remaining': remaining,
                'remaining_currency': remaining_currency
            }

            # Gọi ExcelExporter để tạo báo cáo tháng
            excel_exporter = ExcelExporter(self.group_name)
            _, remaining = excel_exporter.export_daily_balance(data_for_xnk3)

            # Tiếp tục lưu báo cáo ngày như bình thường
            report_dir = f"data/{self.group_name}/reports"
            os.makedirs(report_dir, exist_ok=True)
            
            print(f"Report directory: {report_dir}")  # Debug log

            filename = f"每日报告_{self.group_name}_{date_str}.xlsx"
            filepath = os.path.join(report_dir, filename)
            
            print(f"Saving file to: {filepath}")  # Debug log
            
            try:
                wb.save(filepath)
                print(f"File saved successfully")  # Debug log
                return filepath
            except Exception as e:
                print(f"Error saving file: {str(e)}")  # Debug log
                raise

        except Exception as e:
            print(f"Error creating daily report: {str(e)}")
            traceback.print_exc()
            raise